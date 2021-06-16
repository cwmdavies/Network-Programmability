###############################################
#             Under Construction              #
#                Design Phase                 #
#                                             #
###############################################
#
#   A simple  napalm script that parses the output 
#   of the "show interface descriptions"
#   command and writes it in a neat format
#   to an excel spreadsheet.

import napalm
from openpyxl import load_workbook, Workbook
import datetime as time
import os
from getpass import getpass
import ipaddress

debug = 1
username = input("Please enter your username: ")
password = getpass("Please enter your password: ")
IP_Address = input("Please enter an ip Address: ")


class ExcelWriter:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)

    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames

    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                workbook.create_sheet(value)
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)

    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)

    def filter_cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


#######################################################################################################################
#          Logging Functions
#

def error_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()
    if debug == 1:
        print(message)


def output_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()
    if debug == 1:
        print(message)

#
#
#######################################################################################################################


def main():
    driver_ios = napalm.get_network_driver("ios")    
    device = driver_ios(hostname=IP_Address, username=username, password=password)
    device.open()
    device_interfaces = device.get_interfaces()
    device.close()

    int_detail = ExcelWriter("Interfaces")
    int_detail.add_sheets("Interface configuration",)
    int_detail.write("Interface configuration", "A", "1", "Interface",)
    int_detail.write("Interface configuration", "B", "1", "is_enabled",)
    int_detail.write("Interface configuration", "C", "1", "is_up",)
    int_detail.write("Interface configuration", "D", "1", "Description",)
    int_detail.write("Interface configuration", "E", "1", "MTU",)
    int_detail.write("Interface configuration", "F", "1", "Speed",)
    int_detail.filter_cols("Interface configuration", "A", "30")
    int_detail.filter_cols("Interface configuration", "B", "15")
    int_detail.filter_cols("Interface configuration", "C", "15")
    int_detail.filter_cols("Interface configuration", "D", "50")
    int_detail.filter_cols("Interface configuration", "E", "10")
    int_detail.filter_cols("Interface configuration", "F", "10")

    index = 2
    for interfaces in device_interfaces:
        int_detail.write("Interface configuration", "A", f"{index}", interfaces,)
        int_detail.write("Interface configuration", "B", f"{index}", device_interfaces[interfaces]["is_enabled"],)
        int_detail.write("Interface configuration", "C", f"{index}", device_interfaces[interfaces]["is_up"],)
        int_detail.write("Interface configuration", "D", f"{index}", device_interfaces[interfaces]["description"],)
        int_detail.write("Interface configuration", "E", f"{index}", device_interfaces[interfaces]["mtu"],)
        int_detail.write("Interface configuration", "F", f"{index}", device_interfaces[interfaces]["speed"],)
        index += 1


if __name__ == "__main__":
    main()
