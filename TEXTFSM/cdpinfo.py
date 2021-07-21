###############################################
#            Under Construction               #
#                Design Phase                 #
#                                             #
###############################################

import paramiko
import textfsm
from getpass import getpass
from openpyxl import load_workbook, Workbook
import ipaddress

IPAddr = input("Enter an IP Address: ")
username = input("Enter your username: ")
password = getpass("Enter your Password: ")
jump_server_address = '10.251.6.31'   # The internal ip Address for the Jump server
local_IP_address = '127.0.0.1'  # ip Address of the machine you are connecting from
IP_LIST = []
collection_of_results = []
filename = "CDP_Neighbors_Detail.xlsx"
index = 2


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


# noinspection PyBroadException,PyTypeChecker
def jump_session(ip):
    if not ip_check(ip):
        print(f"open_session function error: "
              f"ip Address {ip} is not a valid Address. Please check and restart the script!", )
        return None, None, False
    try:
        jump_box = paramiko.SSHClient()
        jump_box.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_box.connect(jump_server_address, username=username, password=password)
        jump_box_transport = jump_box.get_transport()
        src_address = (local_IP_address, 22)
        destination_address = (ip, 22)
        jump_box_channel = jump_box_transport.open_channel("direct-tcpip", destination_address, src_address)
        target = paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target.connect(destination_address, username=username, password=password, sock=jump_box_channel)
    except Exception as err:
        print(f"Unable to connect to IP Address: {ip}")
        return None, None, False
    return target, jump_box, True


def get_cdp_details(ip):
    hostname = get_hostname(ip)
    ssh, jump_box, connection = jump_session(ip)
    if not connection:
        return None
    _, stdout, _ = ssh.exec_command("show cdp neighbors detail")
    stdout = stdout.read()
    stdout = stdout.decode("utf-8")
    with open("TEXTFSM_TEMPLATES/cdp_details.txt") as f:
        re_table = textfsm.TextFSM(f)
        result = re_table.ParseText(stdout)
    result = [dict(zip(re_table.header, entry)) for entry in result]
    for entry in result:
        entry['LOCAL_HOST'] = hostname
        collection_of_results.append(entry)

        if 'Switch' in entry['CAPABILITIES']:
            if entry["REMOTE_IP"] not in IP_LIST:
                IP_LIST.append(entry["REMOTE_IP"])

    ssh.close()
    jump_box.close()


def get_hostname(ip):
    ssh, jump_box, connection = jump_session(ip)
    if not connection:
        return None
    _, stdout, _ = ssh.exec_command("show run | inc hostname")
    stdout = stdout.read()
    stdout = stdout.decode("utf-8")
    with open("TEXTFSM_TEMPLATES/hostname.txt") as f:
        re_table = textfsm.TextFSM(f)
        result = re_table.ParseText(stdout)
        hostname = result[0][0]
    ssh.close()
    jump_box.close()
    return hostname


def to_excel(cdp_details):
    global index
    workbook = Workbook()
    workbook.create_sheet("CDP Neighbors Detail")
    del workbook["Sheet"]
    workbook.save(filename=filename)
    workbook = load_workbook(filename=filename)
    ws = workbook["CDP Neighbors Detail"]
    ws["A1"] = "LOCAL_HOST"
    ws["B1"] = "LOCAL_PORT"
    ws["C1"] = "REMOTE_HOST"
    ws["D1"] = "REMOTE_PORT"
    ws["E1"] = "REMOTE_IP"
    ws["F1"] = "PLATFORM"
    ws["G1"] = "SOFTWARE_VERSION"
    ws["H1"] = "CAPABILITIES"
    ws.column_dimensions['A'].width = "25"
    ws.column_dimensions['B'].width = "25"
    ws.column_dimensions['C'].width = "45"
    ws.column_dimensions['D'].width = "25"
    ws.column_dimensions['E'].width = "25"
    ws.column_dimensions['F'].width = "25"
    ws.column_dimensions['G'].width = "120"
    ws.column_dimensions['H'].width = "45"
    workbook.save(filename=filename)
    try:
        for entry in cdp_details:
            ws[f"A{index}"] = entry["LOCAL_HOST"]
            ws[f"B{index}"] = entry["LOCAL_PORT"]
            ws[f"C{index}"] = entry["REMOTE_HOST"]
            ws[f"D{index}"] = entry["REMOTE_PORT"]
            ws[f"E{index}"] = entry["REMOTE_IP"]
            ws[f"F{index}"] = entry["PLATFORM"]
            ws[f"G{index}"] = entry["SOFTWARE_VERSION"]
            ws[f"H{index}"] = entry["CAPABILITIES"]
            workbook.save(filename=filename)
            index += 1

    except Exception as err:
        print("An Exception Occurred")
        print({err})

    workbook.save(filename=filename)


def main():
    IP_LIST.append(IPAddr)

    i = 0
    while i < len(IP_LIST):
        limit = i + min(30, (len(IP_LIST) - i))
        ip_addresses = IP_LIST[i:limit]

        for IP in ip_addresses:
            get_cdp_details(IP)

        i = limit
    to_excel(collection_of_results)


if __name__ == "__main__":
    main()
