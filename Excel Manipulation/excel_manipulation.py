from openpyxl import load_workbook, Workbook


class __excel:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        workbook = Workbook()
        workbook.save(filename=self.filename)
    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames
    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                col_name = workbook.create_sheet(value, self.i)
                self.i += 1
            else:
                print(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)
    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)

book1 = __excel("Test")