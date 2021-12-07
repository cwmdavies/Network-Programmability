###############################################
#            Under Construction               #
#                Design Phase                 #
#                                             #
###############################################

import paramiko
import textfsm
from getpass import getpass
from openpyxl import load_workbook, Workbook
import ipaddress
import logging
import sys


Debugging = 0
IPAddr = input("Enter an IP Address: ")
username = input("Enter your username: ")
password = getpass("Enter your Password: ")
jump_server_address = '10.251.131.6'  # The internal ip Address for the Jump server
local_IP_address = '127.0.0.1'  # ip Address of the machine you are connecting from
IP_LIST = []
Hostnames_List = []
collection_of_results = []
filename = "Interface Details.xlsx"


# ---------------------------------------------------------
# -------------- Logging Configuration Start --------------

# Log file location
logfile = 'debug.log'
# Define the log format
log_format = (
    '[%(asctime)s] %(levelname)-8s %(name)-12s %(message)s')

# Define basic configuration
if Debugging == 0:
    logging.basicConfig(
        # Define logging level
        level=logging.INFO,
        # Declare the object we created to format the log messages
        format=log_format,
        # Declare handlers
        handlers=[
            logging.FileHandler(logfile),
            logging.StreamHandler(sys.stdout),
        ]
    )
elif Debugging == 1:
    logging.basicConfig(
        # Define logging level
        level=logging.DEBUG,
        # Declare the object we created to format the log messages
        format=log_format,
        # Declare handlers
        handlers=[
            logging.FileHandler(logfile),
            logging.StreamHandler(sys.stdout),
        ]
    )

# Define your own logger name
log = logging.getLogger(__name__)

# --------------- Logging Configuration End ---------------
# ---------------------------------------------------------


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


def jump_session(ip):
    if not ip_check(ip):
        log.error(f"open_session function error: "
                  f"ip Address {ip} is not a valid Address. Please check and restart the script!",)
        return None, False
    try:
        log.info(f"Trying to establish a connection to: {ip}")
        jump_box = paramiko.SSHClient()
        jump_box.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_box.connect(jump_server_address, username=username, password=password)
        jump_box_transport = jump_box.get_transport()
        src_address = (local_IP_address, 22)
        destination_address = (ip, 22)
        jump_box_channel = jump_box_transport.open_channel("direct-tcpip", destination_address, src_address, timeout=4,)
        target = paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target.connect(destination_address, username=username, password=password, sock=jump_box_channel, timeout=4,
                       auth_timeout=4, banner_timeout=4)
        log.info(f"Connection to IP: {ip} established")
        return target, jump_box, True
    except paramiko.ssh_exception.AuthenticationException:
        log.error(f"Authentication to IP: {ip} failed! Please check your ip, username and password.")
        return None, None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        log.error(f"Unable to connect to IP: {ip}!")
        return None, None, False
    except (ConnectionError, TimeoutError):
        log.error(f"Connection or Timeout error occurred for IP: {ip}!")
        return None, None, False
    except Exception as err:
        log.error(f"Open Session Error: An unknown error occurred for IP: {ip}!")
        log.error(f"{err}")
        return None, None, False


def get_interfaces(ip):
    ssh, jump_box, connection = jump_session(ip)
    _, stdout, _ = ssh.exec_command("show interfaces")
    stdout = stdout.read()
    stdout = stdout.decode("utf-8")
    with open("cisco_ios_show_interfaces.textfsm") as f:
        re_table = textfsm.TextFSM(f)
        result = re_table.ParseText(stdout)
    results = [dict(zip(re_table.header, entry)) for entry in result]
    ssh.close()
    jump_box.close()
    return results


def main():
    interfaces = get_interfaces(IPAddr)

    index = 2
    workbook = Workbook()
    workbook.create_sheet("Interface Details")
    del workbook["Sheet"]
    workbook.save(filename=filename)
    workbook = load_workbook(filename=filename)
    ws = workbook["Interface Details"]
    ws["A1"] = "INTERFACE"
    ws["B1"] = "LINK_STATUS"
    ws["C1"] = "PROTOCOL_STATUS"
    ws["D1"] = "HARDWARE_TYPE"
    ws["E1"] = "ADDRESS"
    ws["F1"] = "BIA"
    ws["G1"] = "DESCRIPTION"
    ws["H1"] = "IP_ADDRESS"
    ws["I1"] = "MTU"
    ws["J1"] = "DUPLEX"
    ws["K1"] = "SPEED"
    ws["L1"] = "MEDIA_TYPE"
    ws["M1"] = "BANDWIDTH"
    ws["N1"] = "DELAY"
    ws["O1"] = "ENCAPSULATION"
    ws["P1"] = "LAST_INPUT"
    ws["Q1"] = "LAST_OUTPUT"
    ws["R1"] = "LAST_OUTPUT_HANG"
    ws["S1"] = "QUEUE_STRATEGY"
    ws["T1"] = "INPUT_RATE"
    ws["U1"] = "OUTPUT_RATE"
    ws["V1"] = "INPUT_PACKETS"
    ws["W1"] = "OUTPUT_PACKETS"
    ws["X1"] = "INPUT_ERRORS"
    ws["Y1"] = "CRC"
    ws["Z1"] = "ABORT"
    ws["AA1"] = "OUTPUT_ERRORS"
    ws.column_dimensions['A'].width = "30"
    ws.column_dimensions['B'].width = "30"
    ws.column_dimensions['C'].width = "30"
    ws.column_dimensions['D'].width = "30"
    ws.column_dimensions['E'].width = "30"
    ws.column_dimensions['F'].width = "30"
    ws.column_dimensions['G'].width = "50"
    ws.column_dimensions['H'].width = "30"
    ws.column_dimensions['I'].width = "30"
    ws.column_dimensions['J'].width = "30"
    ws.column_dimensions['K'].width = "30"
    ws.column_dimensions['L'].width = "30"
    ws.column_dimensions['M'].width = "30"
    ws.column_dimensions['N'].width = "30"
    ws.column_dimensions['O'].width = "30"
    ws.column_dimensions['P'].width = "30"
    ws.column_dimensions['Q'].width = "30"
    ws.column_dimensions['R'].width = "30"
    ws.column_dimensions['S'].width = "30"
    ws.column_dimensions['T'].width = "30"
    ws.column_dimensions['U'].width = "30"
    ws.column_dimensions['V'].width = "30"
    ws.column_dimensions['W'].width = "30"
    ws.column_dimensions['X'].width = "30"
    ws.column_dimensions['Y'].width = "30"
    ws.column_dimensions['Z'].width = "30"
    ws.column_dimensions['AA'].width = "30"
    workbook.save(filename=filename)

    for entries in interfaces:
        ws[f"A{index}"] = entries["INTERFACE"]
        ws[f"B{index}"] = entries["LINK_STATUS"]
        ws[f"C{index}"] = entries["PROTOCOL_STATUS"]
        ws[f"D{index}"] = entries["HARDWARE_TYPE"]
        ws[f"E{index}"] = entries["ADDRESS"]
        ws[f"F{index}"] = entries["BIA"]
        ws[f"G{index}"] = entries["DESCRIPTION"]
        ws[f"H{index}"] = entries["IP_ADDRESS"]
        ws[f"I{index}"] = entries["MTU"]
        ws[f"J{index}"] = entries["DUPLEX"]
        ws[f"K{index}"] = entries["SPEED"]
        ws[f"L{index}"] = entries["MEDIA_TYPE"]
        ws[f"M{index}"] = entries["BANDWIDTH"]
        ws[f"N{index}"] = entries["DELAY"]
        ws[f"O{index}"] = entries["ENCAPSULATION"]
        ws[f"P{index}"] = entries["LAST_INPUT"]
        ws[f"Q{index}"] = entries["LAST_OUTPUT"]
        ws[f"R{index}"] = entries["LAST_OUTPUT_HANG"]
        ws[f"S{index}"] = entries["QUEUE_STRATEGY"]
        ws[f"T{index}"] = entries["INPUT_RATE"]
        ws[f"U{index}"] = entries["OUTPUT_RATE"]
        ws[f"V{index}"] = entries["INPUT_PACKETS"]
        ws[f"W{index}"] = entries["OUTPUT_PACKETS"]
        ws[f"X{index}"] = entries["INPUT_ERRORS"]
        ws[f"Y{index}"] = entries["CRC"]
        ws[f"Z{index}"] = entries["ABORT"]
        ws[f"AA{index}"] = entries["OUTPUT_ERRORS"]
        index += 1

    workbook.save(filename="Interface Details.xlsx")

    log.info(f"Script Complete", )


if __name__ == "__main__":
    main()
