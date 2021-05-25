###############################################
#             Under Contruction               #
#                Design Phase                 #
#                                             #
###############################################
#
#   A simple script that parses the output 
#   of the "show interface descriptions"
#   command and writes it in a neat format
#   to an excel spreadsheet.

import os
import paramiko
import datetime as time
from openpyxl import load_workbook, Workbook
import re
import time as timer
from getpass import getpass

jumpserver_private_addr = '10.251.6.31'   # The internal IP Address for the Jump server
local_IP_addr = '127.0.0.1' # IP Address of the machine you are connecting from
username = input("Please enter your username: ")
password = getpass("Please enter your password: ")
IP_Addr = input("Please enter an IP Address: ")

interfaces = dict()

class excel_writer:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)
    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames
    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                col_name = workbook.create_sheet(value, self.i)
                self.i += 1
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)
    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)
    def filter_Cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)

def open_session(IP):
    try:
        output_log(f"Trying to establish a connection to: {IP}")
        jumpbox=paramiko.SSHClient()
        jumpbox.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jumpbox.connect(jumpserver_private_addr, username=username, password=password )
        jumpbox_transport = jumpbox.get_transport()
        src_addr = (local_IP_addr, 22)
        dest_addr = (IP, 22)
        jumpbox_channel = jumpbox_transport.open_channel("direct-tcpip", dest_addr, src_addr)
        target=paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target.connect(dest_addr, username=username, password=password, sock=jumpbox_channel)
        output_log(f"Connection to IP: {IP} established")
        return target, jumpbox, True
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Authentication to IP: {IP} failed! Please check your IP, username and password.")
        return None, None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Unable to connect to IP: {IP}!")
        return None, None, False
    except (ConnectionError, TimeoutError):
        error_log(f"Timeout error occured for IP: {IP}!")
        return None, None, False
    except:
        error_log(f"Open Session Error: An unknown error occured for IP: {IP}!")
        return None, None, False

def get_interfaces(IP):
    interface_names = list()
    ssh, jumpbox, connection = open_session(IP)
    if not connection:
        return None
    try:
        output_log(f"retrieving list of interfaces from IP Address: {IP}")
        stdin, stdout, stderr = ssh.exec_command("show ip interface brief")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        regex = r"^(\b(Ten|Gig|Loo|Vla|Fas|Twe|Ten|Fo).{20})"
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        output_log(f"List retrieval successful for IP Address: {IP}")
        return interface_names
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Interfaces function Error: Authentication to IP: {IP} failed! Please check your IP, username and password.")
        return None
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Interfaces function Error: Unable to connect to IP: {IP}!")
        return None
    except (ConnectionError, TimeoutError):
        error_log(f"Interfaces function Error: Timeout error occured for IP: {IP}!")
        return None
    except:
        error_log(f"Interfaces function Error: An unknown error occured for IP: {IP}!")
        return None
    finally:
        ssh.close()
        jumpbox.close()

def get_int_descr(IP, int_name):
    global interfaces
    command = f"show run interface {int_name} | inc description"
    ssh, jumpbox, connection = open_session(IP)
    if not connection:
        error_log(f"get_int_descr - Function Error: No connection is available for IP: {IP}!")
    try:
        output_log(f"retrieving interface description for interface: {int_name}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        Inter_Desc = re.search(".*description.*", stdout)
        Inter_Desc = Inter_Desc[0]
        interfaces[int_name] = Inter_Desc
        output_log(f"Description retrieval successful for interface: {int_name}")
    except TypeError:
        interfaces[int_name] = "No Description found"
    except paramiko.ssh_exception.SSHException:
        error_log(f"get_int_descr - Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
    except:
        error_log(f"get_int_descr - Function Error: An unknown error occured for IP: {IP}, on Interface: {int_name}!")
    finally:
        ssh.close()
        jumpbox.close()


#############################################################################################################################################
##          Logging Functions
#

def error_log(message, debug=0):
    dateTimeObj = time.datetime.now()
    datetime = dateTimeObj.strftime("%d/%m/%Y %H:%M:%S")
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()
    if debug == 1:
        print(message)

def output_log(message, debug=0):
    dateTimeObj = time.datetime.now()
    datetime = dateTimeObj.strftime("%d/%m/%Y %H:%M:%S")
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()
    if debug == 1:
        print(message)

#
##
#############################################################################################################################################

def main():
    global IP_Addr
    global interfaces

    start = timer.time()

    try:
        interface_names = get_interfaces(IP_Addr)

        for int in interface_names:
            get_int_descr(IP_Addr, int)

    except:
        error_log(f"Main function error: An unknown error occured")

    finally:   
        end = timer.time()
        elapsed = (end - start) / 60
        output_log(f"Total execution time: {elapsed:.3} minutes.", i=1)
        output_log(f"Script Complete", i=1)