###############################################
#             Under Contruction               #
#                 Build Phase                 #
#                                             #
###############################################

import os
import re
import time
from multiprocessing.pool import ThreadPool
import paramiko
from datetime import datetime
from getpass import getpass
from openpyxl import load_workbook, Workbook

IP_list = []
CDP_Info_List = []

jumpserver_private_addr = '10.251.6.31'   # The internal ip Address for the Jump server
local_IP_addr = '127.0.0.1' # ip Address of the machine you are connecting from

username = input("Enter your username: ")
password = getpass("Enter your password: ")
IPAddr = input("Enter an ip Address: ")
port = "22"
Sitecode = input("Enter the site code: ")

dateTimeObj = datetime.now()
datetime = dateTimeObj.strftime("%d/%m/%Y %H:%M:%S")

class excel_writer:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)
    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames
    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                col_name = workbook.create_sheet(value, self.i)
                self.i += 1
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)
    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)
    def filter_Cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)

def open_session(IP):
    try:
        output_log(f"Trying to establish a connection to: {IP}")
        jumpbox=paramiko.SSHClient()
        jumpbox.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jumpbox.connect(jumpserver_private_addr, username=username, password=password )
        jumpbox_transport = jumpbox.get_transport()
        src_addr = (local_IP_addr, 22)
        dest_addr = (IP, 22)
        jumpbox_channel = jumpbox_transport.open_channel("direct-tcpip", dest_addr, src_addr)
        target=paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target.connect(dest_addr, username=username, password=password, sock=jumpbox_channel)
        output_log(f"Connection to ip: {IP} established")
        return target, jumpbox, True
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Authentication to ip: {IP} failed! Please check your ip, username and password.")
        return None, None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Unable to connect to ip: {IP}!")
        return None, None, False
    except (ConnectionError, TimeoutError):
        error_log(f"Timeout error occured for ip: {IP}!")
        return None, None, False
    except:
        error_log(f"Open Session Error: An unknown error occured for ip: {IP}!")
        return None, None, False

def extract_cdp_neighbors(IP):
    interface_names = []
    command = "sh cdp neighbors"
    regex = r"^.{17}(\b(Ten|Gig|Loo|Vla|F|Twe|Ten|Fo).{15})"
    ssh, jumpbox, connection = open_session(IP)
    if not connection:
        return None
    try:
        output_log(f"Function Extract CDP Neighbors: Extracting Neighbors: ip Address: {IP}")
        _, output, _ = ssh.exec_command(command)
        output = output.read()
        output = output.decode("utf-8")
        matches = re.finditer(regex, output, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        output_log(f"Function Extract CDP Neighbors: Extraction Complete: ip Address: {IP}")
        return interface_names
    except paramiko.ssh_exception.SSHException:
        error_log(f"Extract CDP Neighbor Function Error: There is an error connecting or establishing SSH session to ip Address {IP}")
        return None, False
    except:
        error_log(f"extract cdp neighbors Error: An unknown error occured for ip: {IP}!")
        return None, False
    finally:
        ssh.close()
        jumpbox.close()

def CDP_Details(IP, commands):
    CDP_Info = {}
    ssh, jumpbox, connection = open_session(IP)
    if not connection:
        return None
    try:
        output_log(f"Function CDP Detail: Extracting Neighbor Details: ip Address: {IP}")
        stdin, stdout, stderr = ssh.exec_command(commands)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")

        RemoteHost = r"(?=[\n\r].*Device ID:[\s]*([^\n\r]*))"
        Platform = r"(?=[\n\r].*Platform:[\s]*([^\n\r]*))"
        Interface = r"(?=[\n\r].*Interface:[\s]*([^\n\r]*))"
        RIPAddr = r"(?=[\n\r].*ip address:[\s]*([^\n\r]*))"
        RemoteInt = r"(?=[\n\r].*Port ID.*: [\s]*([^\n\r]*))"
        Native = r"(?=[\n\r].*Native VLAN:[\s]*([^\n\r]*))"

        RemoteHost_match = re.finditer(RemoteHost, stdout, re.MULTILINE)
        Platform_match = re.finditer(Platform, stdout, re.MULTILINE)
        Interface_match = re.finditer(Interface, stdout, re.MULTILINE)
        RIPAddr_match = re.finditer(RIPAddr, stdout, re.MULTILINE)
        RemoteInt_match = re.finditer(RemoteInt, stdout, re.MULTILINE)
        Native_match = re.finditer(Native, stdout, re.MULTILINE)

        CDP_Info["LocalHost"] = IP

        for line in RemoteHost_match:
            RemoteHost = line[1].split()
            RemoteHost = RemoteHost[0]
            CDP_Info["RemoteHost"] = RemoteHost
        for line in Platform_match:
            Platform = line[1].split(":")
            Platform = line[1].split(",")
            Platform = Platform[0].strip(",")
            CDP_Info["Platform"] = Platform
        for line in Interface_match:
            Interface = line[1].split()
            Interface = Interface[0].strip(",")
            CDP_Info["Local Interface"] = Interface
        for line in RIPAddr_match:
            RIPAddr = line[1].split()
            RIPAddr = RIPAddr[0]
            CDP_Info["Remote ip Address"] = RIPAddr
        for line in RemoteInt_match:
            RemoteInt = line[1].split()
            RemoteInt = RemoteInt[0]
            CDP_Info["Remote Interface"] = RemoteInt
        for line in Native_match:
            Native = line[1].split()
            Native = Native[0]
            CDP_Info["Native VLAN"] = Native
        if RIPAddr not in IP_list:
            IP_list.append(RIPAddr)
        CDP_Info_List.append(CDP_Info)
        output_log(f"Function CDP Detail: Extraction Complete: ip Address: {IP}")
    except paramiko.ssh_exception.SSHException:
        error_log(f"CDP Info Function Error: There is an error connecting or establishing SSH session to ip Address {IP}")
    except:
        error_log(f"CDP Details Error: An unknown error occured for ip: {IP}!")
        return None, False
    finally:
        ssh.close()
        jumpbox.close()

def find_IPs(IP):
    interface_names = extract_cdp_neighbors(IP)
    if not interface_names:
        return -1
    for name in interface_names:
        command = f"show cdp neighbors {name} detail"
        CDP_Details(IP, command)

def error_log(message):
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()

def output_log(message):
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()

def main():
    global IPAddr
    global IP_list
    global CDP_Info_List
    print("Please wait until the script finished - This may take a while depending on the size ofthe network!")
    
    start = time.time()
    IP_list.append(IPAddr)
    pool = ThreadPool(30)
    i = 0

    try:
        while i < len(IP_list):
            limit = i + min(30, (len(IP_list) - i))
            hostnames = IP_list[i:limit]
            pool.map(find_IPs, hostnames)
            i = limit

        pool.close()
        pool.join()

        CDP_Detail = excel_writer(Sitecode)
        CDP_Detail.add_sheets("CDP_Nei_Info",)
        CDP_Detail.write("CDP_Nei_Info","A","1","Local ip Address",)
        CDP_Detail.write("CDP_Nei_Info","B","1","Local Interface",)
        CDP_Detail.write("CDP_Nei_Info","C","1","Remote Interface",)
        CDP_Detail.write("CDP_Nei_Info","D","1","Remote Host",)
        CDP_Detail.write("CDP_Nei_Info","E","1","Remote ip Address",)
        CDP_Detail.write("CDP_Nei_Info","F","1","Platform",)
        CDP_Detail.write("CDP_Nei_Info","G","1","Native VLAN",)
        CDP_Detail.filter_Cols("CDP_Nei_Info","A","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","B","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","C","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","D","45")
        CDP_Detail.filter_Cols("CDP_Nei_Info","E","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","F","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","G","25")

        index = 2
        for entries in CDP_Info_List:
            CDP_Detail.write("CDP_Nei_Info","A",f"{index}",entries["LocalHost"],)
            CDP_Detail.write("CDP_Nei_Info","B",f"{index}",entries["Local Interface"],)
            CDP_Detail.write("CDP_Nei_Info","C",f"{index}",entries["Remote Interface"],)
            CDP_Detail.write("CDP_Nei_Info","D",f"{index}",entries["RemoteHost"],)
            CDP_Detail.write("CDP_Nei_Info","E",f"{index}",entries["Remote ip Address"],)
            CDP_Detail.write("CDP_Nei_Info","F",f"{index}",entries["Platform"],)
            if "Native VLAN" in entries:
                CDP_Detail.write("CDP_Nei_Info","G",f"{index}",entries["Native VLAN"],)
            else:
                CDP_Detail.write("CDP_Nei_Info","G",f"{index}","Not Found",)
            index += 1
    except:
        error_log("Function Main: An unknown error occured!")
    finally:
        end = time.time()
        elapsed = (end - start) / 60
        output_log(f"Total execution time: {elapsed:.3} minutes.")
        output_log("Script Complete!")
        print("Script Complete!")

if __name__ == "__main__":
    main()