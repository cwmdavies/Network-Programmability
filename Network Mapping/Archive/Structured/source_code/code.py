from openpyxl import load_workbook, Workbook
import os
import pandas as pd
import re
from .defaults import *
import napalm

IP_list = []
CDP_Info_List = []
interfaces = []


class ExcelWriter:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)

    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames

    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                workbook.create_sheet(value)
            else:
                print(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)

    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)

    def filter_cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)


def extract_cdp_neighbors(ip):
    interface_names = []
    command = "sh cdp neighbors"
    regex = r"^.{17}(\b(Ten|Gig|Loo|Vla|F|Twe|Ten|Fo).{15})"
    ssh, connection = open_session(ip)
    if not connection:
        return None
    try:
        log.info(f"Extract CDP Neighbors Function: Extracting Neighbors: ip Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        log.info(f"Extract CDP Neighbors Function: Extraction Complete: ip Address: {ip}")
        return interface_names
    except paramiko.ssh_exception.SSHException:
        log.error(f"Extract CDP Neighbors Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
        return None, False
    except Exception as err:
        log.error(f"Extract CDP Neighbors Function Error: An unknown error occurred for ip: {ip}!")
        log.error(f"\t Error: {err}")
        return None, False
    finally:
        ssh.close()


def cdp_details(ip, command, hostname):
    cdp_info = {}
    ssh, connection = open_session(ip)
    if not connection:
        return None
    try:
        log.info(f"CDP Detail Function: Extracting Neighbor Details: ip Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")

        remote_host = r"(?=[\n\r].*Device ID:[\s]*([^\n\r]*))"
        platform = r"(?=[\n\r].*Platform:[\s]*([^\n\r]*))"
        interface = r"(?=[\n\r].*Interface:[\s]*([^\n\r]*))"
        remote_ip_ddr = r"(?=[\n\r].*IP Address:[\s]*([^\n\r]*))"
        remote_int = r"(?=[\n\r].*Port ID.*: [\s]*([^\n\r]*))"
        native_vlan = r"(?=[\n\r].*Native VLAN:[\s]*([^\n\r]*))"

        remote_host_match = re.finditer(remote_host, stdout, re.MULTILINE)
        platform_match = re.finditer(platform, stdout, re.MULTILINE)
        interface_match = re.finditer(interface, stdout, re.MULTILINE)
        remote_ip_addr_match = re.finditer(remote_ip_ddr, stdout, re.MULTILINE)
        remote_int_match = re.finditer(remote_int, stdout, re.MULTILINE)
        native_vlan_match = re.finditer(native_vlan, stdout, re.MULTILINE)

        cdp_info["Local Hostname"] = hostname
        cdp_info["Local ip Address"] = ip

        for line in remote_host_match:
            remote_host = line[1].split()
            remote_host = remote_host[0]
            cdp_info["Remote Host"] = remote_host
        for line in platform_match:
            platform = line[1].split(",")
            platform = platform[0].strip(",")
            cdp_info["Platform"] = platform
        for line in interface_match:
            interface = line[1].split()
            interface = interface[0].strip(",")
            cdp_info["Local Interface"] = interface
        for line in remote_ip_addr_match:
            remote_ip_ddr = line[1].split()
            remote_ip_ddr = remote_ip_ddr[0]
            cdp_info["Remote ip Address"] = remote_ip_ddr
        for line in remote_int_match:
            remote_int = line[1].split(":")
            remote_int = remote_int[0]
            cdp_info["Remote Interface"] = remote_int
        for line in native_vlan_match:
            native_vlan = line[1].split()
            native_vlan = native_vlan[0]
            cdp_info["Native VLAN"] = native_vlan
        if remote_ip_ddr not in IP_list:
            IP_list.append(remote_ip_ddr)
        CDP_Info_List.append(cdp_info)
        log.info(f"CDP Detail Function: Extraction Complete: ip Address: {ip}")
    except paramiko.ssh_exception.SSHException:
        log.error(f"CDP Detail Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
    except Exception as err:
        log.error(f"CDP Details Function Error: An unknown error occurred for ip: {ip}!")
        log.error(f"\t Error: {err}")
        return None, False
    finally:
        ssh.close()


def get_hostname(ip):
    hostname = None
    regex_hostname = r"^\bhostname[\s\r]+(.*)$"
    ssh, connection = open_session(ip)
    if not connection:
        return "-1"
    try:
        log.info(f"Get Hostname Function: Extracting Hostname: IP Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command("show run | i hostname")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        hostname_matches = re.finditer(regex_hostname, stdout, re.MULTILINE)
        for h in hostname_matches:
            hostname = h.group(1)
        return hostname
    except paramiko.ssh_exception.SSHException:
        log.error(f"Get Hostname Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
        return None
    except Exception as err:
        log.error(f"Get Hostname Function Error: Unknown error occurred for ip Address: {ip}!")
        log.error(f"\t Error: {err}")
        return None
    finally:
        ssh.close()


def find_ips(ip):
    interface_names = extract_cdp_neighbors(ip)
    hostname = get_hostname(ip)
    if not interface_names:
        return -1
    for name in interface_names:
        command = f"show cdp neighbors {name} detail"
        cdp_details(ip, command, hostname)


def get_interfaces(ip):
    interface_names = list()
    ssh, jump_box, connection = jump_session(ip)
    if not connection:
        return None
    try:
        log.info(f"retrieving list of interfaces from ip Address: {ip}")
        _, stdout, _ = ssh.exec_command("show ip interface brief")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        regex = r"^(\b(Ten|Gig|Loo|Vla|Fas|Twe|Ten|Fo).{20})"
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        log.info(f"List retrieval successful for ip Address: {ip}")
        return interface_names
    except paramiko.ssh_exception.AuthenticationException:
        log.error(f"Interfaces function Error: Authentication to ip: "
                  f"{ip} failed! Please check your ip, username and password.")
        return None
    except paramiko.ssh_exception.NoValidConnectionsError:
        log.error(f"Interfaces function Error: Unable to connect to ip: {ip}!")
        return None
    except (ConnectionError, TimeoutError):
        log.error(f"Interfaces function Error: Timeout error occurred for ip: {ip}!")
        return None
    except Exception as err:
        log.error(f"Interfaces function Error: An unknown error occurred for ip: {ip}!")
        log.error(f"\t Error: {err}")
        return None
    finally:
        ssh.close()
        jump_box.close()


def get_int_description(int_name):
    global interfaces
    interfaces_dict = dict()
    command = f"show run interface {int_name} | inc description"
    ssh, jump_box, connection = jump_session(IP_Address)
    if not connection:
        log.error(f"get_int_description - Function Error: No connection is available for ip: {IP_Address}!")
    try:
        log.info(f"retrieving interface description for interface: {int_name}")
        _, stdout, _ = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        int_description = re.search(".*description.*", stdout)
        int_description = int_description[0]
        int_description = int_description.strip()
        int_description = int_description.strip("description")
        interfaces_dict["Interface"] = int_name
        interfaces_dict["Description"] = int_description
        log.info(f"Description retrieval successful for interface: {int_name}")
    except TypeError:
        interfaces_dict["Interface"] = int_name
        interfaces_dict["Description"] = "No Description found"
    except paramiko.ssh_exception.SSHException:
        log.error(f"get_int_description - Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {IP_Address}")
    except Exception as err:
        log.error(f"get_int_description - Function Error: An unknown error occurred for ip: {IP_Address}, "
                  f"on Interface: {int_name}!")
        log.error(f"\t Error: {err}")
    finally:
        interfaces.append(interfaces_dict)
        ssh.close()
        jump_box.close()


def int_write(ip):
    df = pd.read_excel(r'Interfaces.xlsx')
    commands = []
    ssh, jump_box, connection = jump_session(ip)
    if not connection:
        return None
    try:
        log.info(f"int_write function: Preparing to writing interface descriptions.")
        channel = ssh.invoke_shell()
        stdin = channel.makefile("wb")
        output = channel.makefile("rb")
        commands.append("'''")
        commands.append("conf t")
        for num in range(len(df)):
            commands.append(f"interface {df['Interface'][num]}")
            commands.append(f"description {df['Description'][num]}")
        commands.append("end")
        commands.append("exit")
        commands.append("'''")
        commands = "\n".join(commands)
        stdin.write(str.encode(commands))
        output = output.read()
        output = output.decode("utf-8")
        log.info(f"Output:\n\t{output}")
        stdin.close()
        log.info(f"int_write function: Finished writing interface descriptions.")
    except Exception as err:
        log.error(f"Int_write function Error: An unknown error occurred!")
        log.error(f"\t Error: {err}")
    finally:
        ssh.close()
        jump_box.close()


def np_get_interfaces(ip):
    driver_ios = napalm.get_network_driver("ios")
    device = driver_ios(hostname=ip, username=username, password=password)
    device.open()
    device_interfaces = device.get_interfaces()
    device.close()
    return device_interfaces


def np_get_hostname(ip):
    driver_ios = napalm.get_network_driver("ios")
    device = driver_ios(hostname=ip, username=username, password=password)
    device.open()
    device_facts = device.get_facts()
    device.close()
    return device_facts["hostname"]