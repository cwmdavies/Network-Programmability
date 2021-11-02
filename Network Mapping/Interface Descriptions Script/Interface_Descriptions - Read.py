###############################################
#             Under Construction               #
#               Testing Phase                 #
#                                             #
###############################################
#
#   A simple script that parses the output 
#   of the "show interface descriptions"
#   command and writes it in a neat format
#   to an excel spreadsheet.

import paramiko
from openpyxl import load_workbook, Workbook
import re
import time as timer
from getpass import getpass
import ipaddress
import logging
import sys

Debugging = 0
jump_server_address = '10.251.6.31'   # The internal ip Address for the Jump server
local_IP_address = '127.0.0.1'  # ip Address of the machine you are connecting from
username = input("Please enter your username: ")
password = getpass("Please enter your password: ")
IP_Address = input("Please enter an ip Address: ")
debug = 0
interfaces = list()
filename = "Interface Descriptions.xlsx"

# ---------------------------------------------------------
# -------------- Logging Configuration Start --------------

# Log file location
logfile = 'debug.log'
# Define the log format
log_format = (
    '[%(asctime)s] %(levelname)-8s %(name)-12s %(message)s')

# Define basic configuration
if Debugging == 0:
    logging.basicConfig(
        # Define logging level
        level=logging.INFO,
        # Declare the object we created to format the log messages
        format=log_format,
        # Declare handlers
        handlers=[
            logging.FileHandler(logfile),
            logging.StreamHandler(sys.stdout),
        ]
    )
elif Debugging == 1:
    logging.basicConfig(
        # Define logging level
        level=logging.DEBUG,
        # Declare the object we created to format the log messages
        format=log_format,
        # Declare handlers
        handlers=[
            logging.FileHandler(logfile),
            logging.StreamHandler(sys.stdout),
        ]
    )

# Define your own logger name
log = logging.getLogger(__name__)

# --------------- Logging Configuration End ---------------
# ---------------------------------------------------------


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


def open_session(ip):
    if not ip_check(ip):
        log.error(f"open_session function error: "
                  f"ip Address {ip} is not a valid Address. Please check and restart the script!",)
        return None, False
    try:
        log.info(f"Trying to establish a connection to: {ip}")
        jump_box = paramiko.SSHClient()
        jump_box.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_box.connect(jump_server_address, username=username, password=password)
        jump_box_transport = jump_box.get_transport()
        src_address = (local_IP_address, 22)
        destination_address = (ip, 22)
        jump_box_channel = jump_box_transport.open_channel("direct-tcpip", destination_address, src_address, timeout=4,)
        target = paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target.connect(destination_address, username=username, password=password, sock=jump_box_channel, timeout=4,
                       auth_timeout=4, banner_timeout=4)
        log.info(f"Connection to ip: {ip} established")
        return target, jump_box, True
    except paramiko.ssh_exception.AuthenticationException:
        log.error(f"Authentication to ip: {ip} failed! Please check your ip, username and password.")
        return None, None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        log.error(f"Unable to connect to ip: {ip}!")
        return None, None, False
    except (ConnectionError, TimeoutError):
        log.error(f"Timeout error occurred for ip: {ip}!")
        return None, None, False
    except Exception as err:
        log.error(f"Open Session Error: An unknown error occurred for ip: {ip}!")
        log.error(f"\t Error: {err}")
        return None, None, False


def get_interfaces(ip):
    interface_names = list()
    ssh, jump_box, connection = open_session(ip)
    if not connection:
        return None
    try:
        log.info(f"retrieving list of interfaces from ip Address: {ip}")
        _, stdout, _ = ssh.exec_command("show ip interface brief")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        regex = r"^(\b(Ten|Gig|Loo|Vla|Fas|Twe|Ten|Fo).{20})"
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        log.info(f"List retrieval successful for ip Address: {ip}")
        return interface_names
    except paramiko.ssh_exception.AuthenticationException:
        log.error(f"Interfaces function Error: Authentication to ip: "
                  f"{ip} failed! Please check your ip, username and password.")
        return None
    except paramiko.ssh_exception.NoValidConnectionsError:
        log.error(f"Interfaces function Error: Unable to connect to ip: {ip}!")
        return None
    except (ConnectionError, TimeoutError):
        log.error(f"Interfaces function Error: Timeout error occurred for ip: {ip}!")
        return None
    except Exception as err:
        log.error(f"Interfaces function Error: An unknown error occurred for ip: {ip}!")
        log.error(f"\t Error: {err}")
        return None
    finally:
        ssh.close()
        jump_box.close()


def get_int_description(int_name):
    global interfaces
    interfaces_dict = dict()
    command = f"show run interface {int_name} | inc description"
    ssh, jump_box, connection = open_session(IP_Address)
    if not connection:
        log.error(f"get_int_description - Function Error: No connection is available for ip: {IP_Address}!")
    try:
        log.info(f"retrieving interface description for interface: {int_name}")
        _, stdout, _ = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        int_description = re.search(".*description.*", stdout)
        int_description = int_description[0]
        int_description = int_description.strip()
        int_description = int_description.strip("description")
        interfaces_dict["Interface"] = int_name
        interfaces_dict["Description"] = int_description
        log.info(f"Description retrieval successful for interface: {int_name}")
    except TypeError:
        interfaces_dict["Interface"] = int_name
        interfaces_dict["Description"] = "No Description found"
    except paramiko.ssh_exception.SSHException:
        log.error(f"get_int_description - Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {IP_Address}")
    except Exception as err:
        log.error(f"get_int_description - Function Error: An unknown error occurred for ip: {IP_Address}, "
                  f"on Interface: {int_name}!")
        log.error(f"\t Error: {err}")
    finally:
        interfaces.append(interfaces_dict)
        ssh.close()
        jump_box.close()


def main():
    global IP_Address
    global interfaces
    
    start = timer.time()

    try:
        interface_names = get_interfaces(IP_Address)

        for int_name in interface_names:
            get_int_description(int_name)

        index = 2
        workbook = Workbook()
        workbook.create_sheet("Interface Descriptions")
        del workbook["Sheet"]
        workbook.save(filename=filename)
        workbook = load_workbook(filename=filename)
        ws = workbook["Interface Descriptions"]
        ws["A1"] = "Interface"
        ws["B1"] = "Description"
        ws.column_dimensions['A'].width = "30"
        ws.column_dimensions['B'].width = "30"
        workbook.save(filename=filename)

        for entries in interfaces:
            ws[f"A{index}"] = entries["Interface"]
            ws[f"B{index}"] = entries["Description"]
            index += 1

            workbook.save(filename=filename)

    finally:   
        end = timer.time()
        elapsed = (end - start) / 60
        log.info(f"Total execution time: {elapsed:.3} minutes.",)
        log.info(f"Script Complete",)


if __name__ == "__main__":
    main()
