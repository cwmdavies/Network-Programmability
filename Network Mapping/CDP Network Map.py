###############################################
#             Under Contruction               #
#               Testing Phase                 #
#                                             #
###############################################

import os
import re
import time as timer
from multiprocessing.pool import ThreadPool
import paramiko
import datetime as time
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox

IP_list = []
CDP_Info_List = []
port = "22"

class excel_writer:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)
    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames
    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                col_name = workbook.create_sheet(value, self.i)
                self.i += 1
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)
    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)
    def filter_Cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)

############################# Start of Tkinter Code #############################

# root window
root = tk.Tk()
root.eval('tk::PlaceWindow . center')
root.geometry("300x250")
root.resizable(False, False)
root.title('Site Details')

# store entries
Username_var = tk.StringVar()
password_var = tk.StringVar()
IP_Address_var = tk.StringVar()
Site_code_var = tk.StringVar()

# Site details frame
Site_details = ttk.Frame(root)
Site_details.pack(padx=10, pady=10, fill='x', expand=True)

# Username
Username_label = ttk.Label(Site_details, text="Username:")
Username_label.pack(fill='x', expand=True)

Username_entry = ttk.Entry(Site_details, textvariable=Username_var)
Username_entry.pack(fill='x', expand=True)
Username_entry.focus()

# Password
password_label = ttk.Label(Site_details, text="Password:" )
password_label.pack(fill='x', expand=True)

password_entry = ttk.Entry(Site_details, textvariable=password_var, show="*")
password_entry.pack(fill='x', expand=True)

# IP Address
IP_Address_label = ttk.Label(Site_details, text="IP Address:")
IP_Address_label.pack(fill='x', expand=True)

IP_Address_entry = ttk.Entry(Site_details, textvariable=IP_Address_var)
IP_Address_entry.pack(fill='x', expand=True)

# Site Code
Site_code_label = ttk.Label(Site_details, text="Site code:")
Site_code_label.pack(fill='x', expand=True)

Site_code_entry = ttk.Entry(Site_details, textvariable=Site_code_var)
Site_code_entry.pack(fill='x', expand=True)

# Submit button
Submit_button = ttk.Button(Site_details, text="Submit", command=root.destroy)
Submit_button.pack(fill='x', expand=True, pady=10)

root.attributes('-topmost', True)
root.mainloop()

username = Username_var.get()
password = password_var.get()
IPAddr = IP_Address_var.get()
Sitecode = Site_code_var.get()

def MessageBox(text, title):
        root = tkinter.Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        tkinter.messagebox.showinfo(title, text)
        root.destroy()

############################# End of Tkinter Code #############################

def open_session(IP):
    try:
        output_log(f"Open Session Function: Trying to connect to IP Address: {IP}")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=IP, port=port, username=username, password=password)
        output_log(f"Open Session Function: Connected to IP Address: {IP}")
        return ssh, True
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Open Session Function: Authentication to IP Address: {IP} failed! Please check your IP, username and password.")
        return None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Open Session Function Error: Unable to connect to IP Address: {IP}!")
        return None, False
    except (ConnectionError, TimeoutError):
        error_log(f"Open Session Function Error: Timeout error occured for IP Address: {IP}!")
        return None, False

def extract_cdp_neighbors(IP):
    interface_names = []
    command = "sh cdp neighbors"
    regex = r"^.{17}(\b(Ten|Gig|Loo|Vla|F|Twe|Ten|Fo).{15})"
    ssh, connection = open_session(IP)
    if not connection:
        return None
    try:
        output_log(f"Extract CDP Neighbors Function: Extracting Neighbors: IP Address: {IP}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        output_log(f"Extract CDP Neighbors Function: Extraction Complete: IP Address: {IP}")
        return interface_names
    except paramiko.ssh_exception.SSHException:
        error_log(f"Extract CDP Neighbors Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
        return None, False
    except:
        error_log(f"Extract CDP Neighbors Function Error: An unknown error occured for IP: {IP}!")
        return None, False
    finally:
        ssh.close()

def CDP_Details(IP, command, hostname):
    CDP_Info = {}
    ssh, connection = open_session(IP)
    if not connection:
        return None
    try:
        output_log(f"CDP Detail Function: Extracting Neighbor Details: IP Address: {IP}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")

        RemoteHost = r"(?=[\n\r].*Device ID:[\s]*([^\n\r]*))"
        Platform = r"(?=[\n\r].*Platform:[\s]*([^\n\r]*))"
        Interface = r"(?=[\n\r].*Interface:[\s]*([^\n\r]*))"
        RIPAddr = r"(?=[\n\r].*IP address:[\s]*([^\n\r]*))"
        RemoteInt = r"(?=[\n\r].*Port ID.*: [\s]*([^\n\r]*))"
        Native = r"(?=[\n\r].*Native VLAN:[\s]*([^\n\r]*))"

        RemoteHost_match = re.finditer(RemoteHost, stdout, re.MULTILINE)
        Platform_match = re.finditer(Platform, stdout, re.MULTILINE)
        Interface_match = re.finditer(Interface, stdout, re.MULTILINE)
        RIPAddr_match = re.finditer(RIPAddr, stdout, re.MULTILINE)
        RemoteInt_match = re.finditer(RemoteInt, stdout, re.MULTILINE)
        Native_match = re.finditer(Native, stdout, re.MULTILINE)

        CDP_Info["Local Hostname"] = hostname
        CDP_Info["Local IP Address"] = IP

        for line in RemoteHost_match:
            RemoteHost = line[1].split()
            RemoteHost = RemoteHost[0]
            CDP_Info["Remote Host"] = RemoteHost
        for line in Platform_match:
            Platform = line[1].split(":")
            Platform = line[1].split(",")
            Platform = Platform[0].strip(",")
            CDP_Info["Platform"] = Platform
        for line in Interface_match:
            Interface = line[1].split()
            Interface = Interface[0].strip(",")
            CDP_Info["Local Interface"] = Interface
        for line in RIPAddr_match:
            RIPAddr = line[1].split()
            RIPAddr = RIPAddr[0]
            CDP_Info["Remote IP Address"] = RIPAddr
        for line in RemoteInt_match:
            RemoteInt = line[1].split(":")
            RemoteInt = RemoteInt[0]
            CDP_Info["Remote Interface"] = RemoteInt
        for line in Native_match:
            Native = line[1].split()
            Native = Native[0]
            CDP_Info["Native VLAN"] = Native
        if RIPAddr not in IP_list:
            IP_list.append(RIPAddr)
        CDP_Info_List.append(CDP_Info)
        output_log(f"CDP Detail Function: Extraction Complete: IP Address: {IP}")
    except paramiko.ssh_exception.SSHException:
        error_log(f"CDP Detail Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
    except:
        error_log(f"CDP Details Function Error: An unknown error occured for IP: {IP}!")
        return None, False
    finally:
        ssh.close()

def get_hostname(IP):
    hostname = None
    regex_hostname = r"^\bhostname[\s\r]+(.*)$"
    ssh, connection = open_session(IP)
    if not connection:
        return "-1"
    try:
        output_log(f"Get Hostname Function: Extracting Hostname: IP Address: {IP}")        
        stdin, stdout, stderr = ssh.exec_command("show run | i hostname")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        hostname_matches = re.finditer(regex_hostname, stdout, re.MULTILINE)
        for h in hostname_matches:
            hostname = h.group(1)
        return hostname
    except paramiko.ssh_exception.SSHException:
        error_log(f"Get Hostname Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
    finally:
        ssh.close()

def find_IPs(IP):
    interface_names = extract_cdp_neighbors(IP)
    hostname = get_hostname(IP)
    if not interface_names:
        return -1
    for name in interface_names:
        command = f"show cdp neighbors {name} detail"
        CDP_Details(IP, command, hostname)

def error_log(message, i=0):
    dateTimeObj = time.datetime.now()
    datetime = dateTimeObj.strftime("%d/%m/%Y %H:%M:%S")
    error_file = open(f"{Sitecode} - Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()
    if i == 1:
        print(message)

def output_log(message, i=0):
    dateTimeObj = time.datetime.now()
    datetime = dateTimeObj.strftime("%d/%m/%Y %H:%M:%S")
    output_file = open(f"{Sitecode} - Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()
    if i == 1:
        print(message)

def main():
    global CDP_Info_List
    global IP_list

    start = timer.time()
    IP_list.append(IPAddr)
    pool = ThreadPool(30)
    i = 0

    try:
        output_log(f"Script started for site: {Sitecode}", i=1)
        print("You will be notified when the script finishes - This may take a while depending on the size of the network!")
        
        while i < len(IP_list):
            limit = i + min(30, (len(IP_list) - i))
            hostnames = IP_list[i:limit]
            pool.map(find_IPs, hostnames)
            i = limit

        pool.close()
        pool.join()

        CDP_Detail = excel_writer(Sitecode)
        CDP_Detail.add_sheets("CDP_Nei_Info",)
        CDP_Detail.write("CDP_Nei_Info","A","1","Local Hostname",)
        CDP_Detail.write("CDP_Nei_Info","B","1","Local IP Address",)
        CDP_Detail.write("CDP_Nei_Info","C","1","Local Interface",)
        CDP_Detail.write("CDP_Nei_Info","D","1","Remote Interface",)
        CDP_Detail.write("CDP_Nei_Info","E","1","Remote Hostname",)
        CDP_Detail.write("CDP_Nei_Info","F","1","Remote IP Address",)
        CDP_Detail.write("CDP_Nei_Info","G","1","Platform",)
        CDP_Detail.write("CDP_Nei_Info","H","1","Native VLAN",)
        CDP_Detail.filter_Cols("CDP_Nei_Info","A","30")
        CDP_Detail.filter_Cols("CDP_Nei_Info","B","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","C","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","D","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","E","45")
        CDP_Detail.filter_Cols("CDP_Nei_Info","F","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","G","25")
        CDP_Detail.filter_Cols("CDP_Nei_Info","H","25")

        index = 2
        for entries in CDP_Info_List:
            CDP_Detail.write("CDP_Nei_Info","A",f"{index}",entries["Local Hostname"],)
            CDP_Detail.write("CDP_Nei_Info","B",f"{index}",entries["Local IP Address"],)
            CDP_Detail.write("CDP_Nei_Info","C",f"{index}",entries["Local Interface"],)
            CDP_Detail.write("CDP_Nei_Info","D",f"{index}",entries["Remote Interface"],)
            CDP_Detail.write("CDP_Nei_Info","E",f"{index}",entries["Remote Host"],)
            CDP_Detail.write("CDP_Nei_Info","F",f"{index}",entries["Remote IP Address"],)
            CDP_Detail.write("CDP_Nei_Info","G",f"{index}",entries["Platform"],)
            if "Native VLAN" in entries:
                CDP_Detail.write("CDP_Nei_Info","H",f"{index}",entries["Native VLAN"],)
            else:
                CDP_Detail.write("CDP_Nei_Info","H",f"{index}","Not Found",)
            index += 1
    except:
        error_log("Main Function Error: An unknown error occured!")
    finally:
        end = timer.time()
        elapsed = (end - start) / 60
        output_log(f"Total execution time: {elapsed:.3} minutes.", i=1)
        output_log(f"Script Complete for site: {Sitecode}", i=1)
        MessageBox(f"Script Complete for site: {Sitecode}", "Script Complete")

if __name__ == "__main__":
    main()