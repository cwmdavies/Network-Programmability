###############################################
#            Under Construction               #
#               Testing Phase                 #
#                                             #
###############################################

import os
import re
import time as timer
from multiprocessing.pool import ThreadPool
import paramiko
import datetime as time
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox
import ipaddress

IP_list = []
CDP_Info_List = []
debug = 0


class ExcelWriter:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        if os.path.exists(f"{self.filename}"):
            os.remove(f"{self.filename}")
        workbook = Workbook()
        workbook.save(filename=self.filename)

    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames

    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                workbook.create_sheet(value)
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)

    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)

    def filter_cols(self, sheet, col, width):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions[f'{col}'].width = width
        workbook.save(filename=self.filename)

##########################################################
# Start of Tkinter Code


# root window
root = tk.Tk()
root.eval('tk::PlaceWindow . center')
root.geometry("300x250")
root.resizable(False, False)
root.title('Site Details')


# store entries
Username_var = tk.StringVar()
password_var = tk.StringVar()
IP_Address_var = tk.StringVar()
Site_code_var = tk.StringVar()


# Site details frame
Site_details = ttk.Frame(root)
Site_details.pack(padx=10, pady=10, fill='x', expand=True)


# Username
Username_label = ttk.Label(Site_details, text="Username:")
Username_label.pack(fill='x', expand=True)


Username_entry = ttk.Entry(Site_details, textvariable=Username_var)
Username_entry.pack(fill='x', expand=True)
Username_entry.focus()


# Password
password_label = ttk.Label(Site_details, text="Password:")
password_label.pack(fill='x', expand=True)
password_entry = ttk.Entry(Site_details, textvariable=password_var, show="*")
password_entry.pack(fill='x', expand=True)


# ip Address
IP_Address_label = ttk.Label(Site_details, text="ip Address:")
IP_Address_label.pack(fill='x', expand=True)
IP_Address_entry = ttk.Entry(Site_details, textvariable=IP_Address_var)
IP_Address_entry.pack(fill='x', expand=True)


# Site Code
Site_code_label = ttk.Label(Site_details, text="Site code:")
Site_code_label.pack(fill='x', expand=True)
Site_code_entry = ttk.Entry(Site_details, textvariable=Site_code_var)
Site_code_entry.pack(fill='x', expand=True)


# Submit button
Submit_button = ttk.Button(Site_details, text="Submit", command=root.destroy)
Submit_button.pack(fill='x', expand=True, pady=10)


root.attributes('-topmost', True)
root.mainloop()


username = Username_var.get()
password = password_var.get()
IPAddr = IP_Address_var.get()
Sitecode = Site_code_var.get()


def messagebox(text, title):
    message = tkinter.Tk()
    message.attributes('-topmost', True)
    message.withdraw()
    tkinter.messagebox.showinfo(title, text)
    message.destroy()

# End of Tkinter Code
##########################################################


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


def open_session(ip):
    if not ip_check(ip):
        return None, False
    try:
        output_log(f"Open Session Function: Trying to connect to ip Address: {ip}")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=ip, port=22, username=username, password=password)
        output_log(f"Open Session Function: Connected to ip Address: {ip}")
        return ssh, True
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Open Session Function:"
                  f"Authentication to ip Address: {ip} failed! Please check your ip, username and password.")
        return None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Open Session Function Error: Unable to connect to ip Address: {ip}!")
        return None, False
    except (ConnectionError, TimeoutError):
        error_log(f"Open Session Function Error: Timeout error occurred for ip Address: {ip}!")
        return None, False
    except Exception as err:
        error_log(f"Open Session Function Error: Unknown error occurred for ip Address: {ip}!")
        error_log(f"\t Error: {err}")
        return None, False


def extract_cdp_neighbors(ip):
    interface_names = []
    command = "sh cdp neighbors"
    regex = r"^.{17}(\b(Ten|Gig|Loo|Vla|F|Twe|Ten|Fo).{15})"
    ssh, connection = open_session(ip)
    if not connection:
        return None
    try:
        output_log(f"Extract CDP Neighbors Function: Extracting Neighbors: ip Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        output_log(f"Extract CDP Neighbors Function: Extraction Complete: ip Address: {ip}")
        return interface_names
    except paramiko.ssh_exception.SSHException:
        error_log(f"Extract CDP Neighbors Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
        return None, False
    except Exception as err:
        error_log(f"Extract CDP Neighbors Function Error: An unknown error occurred for ip: {ip}!")
        error_log(f"\t Error: {err}")
        return None, False
    finally:
        ssh.close()


def cdp_details(ip, command, hostname):
    cdp_info = {}
    ssh, connection = open_session(ip)
    if not connection:
        return None
    try:
        output_log(f"CDP Detail Function: Extracting Neighbor Details: ip Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")

        remote_host = r"(?=[\n\r].*Device ID:[\s]*([^\n\r]*))"
        platform = r"(?=[\n\r].*Platform:[\s]*([^\n\r]*))"
        interface = r"(?=[\n\r].*Interface:[\s]*([^\n\r]*))"
        remote_ip_ddr = r"(?=[\n\r].*IP Address:[\s]*([^\n\r]*))"
        remote_int = r"(?=[\n\r].*Port ID.*: [\s]*([^\n\r]*))"
        native_vlan = r"(?=[\n\r].*Native VLAN:[\s]*([^\n\r]*))"

        remote_host_match = re.finditer(remote_host, stdout, re.MULTILINE)
        platform_match = re.finditer(platform, stdout, re.MULTILINE)
        interface_match = re.finditer(interface, stdout, re.MULTILINE)
        remote_ip_addr_match = re.finditer(remote_ip_ddr, stdout, re.MULTILINE)
        remote_int_match = re.finditer(remote_int, stdout, re.MULTILINE)
        native_vlan_match = re.finditer(native_vlan, stdout, re.MULTILINE)

        cdp_info["Local Hostname"] = hostname
        cdp_info["Local ip Address"] = ip

        for line in remote_host_match:
            remote_host = line[1].split()
            remote_host = remote_host[0]
            cdp_info["Remote Host"] = remote_host
        for line in platform_match:
            platform = line[1].split(",")
            platform = platform[0].strip(",")
            cdp_info["Platform"] = platform
        for line in interface_match:
            interface = line[1].split()
            interface = interface[0].strip(",")
            cdp_info["Local Interface"] = interface
        for line in remote_ip_addr_match:
            remote_ip_ddr = line[1].split()
            remote_ip_ddr = remote_ip_ddr[0]
            cdp_info["Remote ip Address"] = remote_ip_ddr
        for line in remote_int_match:
            remote_int = line[1].split(":")
            remote_int = remote_int[0]
            cdp_info["Remote Interface"] = remote_int
        for line in native_vlan_match:
            native_vlan = line[1].split()
            native_vlan = native_vlan[0]
            cdp_info["Native VLAN"] = native_vlan
        if remote_ip_ddr not in IP_list:
            IP_list.append(remote_ip_ddr)
        CDP_Info_List.append(cdp_info)
        output_log(f"CDP Detail Function: Extraction Complete: ip Address: {ip}")
    except paramiko.ssh_exception.SSHException:
        error_log(f"CDP Detail Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
    except Exception as err:
        error_log(f"CDP Details Function Error: An unknown error occurred for ip: {ip}!")
        error_log(f"\t Error: {err}")
        return None, False
    finally:
        ssh.close()


def get_hostname(ip):
    hostname = None
    regex_hostname = r"^\bhostname[\s\r]+(.*)$"
    ssh, connection = open_session(ip)
    if not connection:
        return "-1"
    try:
        output_log(f"Get Hostname Function: Extracting Hostname: IP Address: {ip}")
        stdin, stdout, stderr = ssh.exec_command("show run | i hostname")
        stdout = stdout.read()
        stdout = stdout.decode("utf-8")
        hostname_matches = re.finditer(regex_hostname, stdout, re.MULTILINE)
        for h in hostname_matches:
            hostname = h.group(1)
        return hostname
    except paramiko.ssh_exception.SSHException:
        error_log(f"Get Hostname Function Error: "
                  f"There is an error connecting or establishing SSH session to ip Address {ip}")
        return None
    except Exception as err:
        error_log(f"Get Hostname Function Error: Unknown error occurred for ip Address: {ip}!")
        error_log(f"\t Error: {err}")
        return None
    finally:
        ssh.close()


def find_ips(ip):
    interface_names = extract_cdp_neighbors(ip)
    hostname = get_hostname(ip)
    if not interface_names:
        return -1
    for name in interface_names:
        command = f"show cdp neighbors {name} detail"
        cdp_details(ip, command, hostname)


#######################################################################################################################
#          Logging Functions
#

def error_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()
    if debug == 1:
        print(message)


def output_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()
    if debug == 1:
        print(message)

#
#
#######################################################################################################################


def main():
    global CDP_Info_List
    global IP_list

    start = timer.time()
    IP_list.append(IPAddr)
    pool = ThreadPool(30)
    i = 0

    try:
        output_log(f"Script started for site: {Sitecode}",)
        print("You will be notified when the script finishes - "
              "This may take a while depending on the size of the network!")
        
        while i < len(IP_list):
            limit = i + min(30, (len(IP_list) - i))
            hostnames = IP_list[i:limit]
            pool.map(find_ips, hostnames)
            i = limit

        pool.close()
        pool.join()

        cdp_detail = ExcelWriter(Sitecode)
        cdp_detail.add_sheets("CDP_Nei_Info",)
        cdp_detail.write("CDP_Nei_Info", "A", "1", "Local Hostname",)
        cdp_detail.write("CDP_Nei_Info", "B", "1", "Local ip Address",)
        cdp_detail.write("CDP_Nei_Info", "C", "1", "Local Interface",)
        cdp_detail.write("CDP_Nei_Info", "D", "1", "Remote Interface",)
        cdp_detail.write("CDP_Nei_Info", "E", "1", "Remote Hostname",)
        cdp_detail.write("CDP_Nei_Info", "F", "1", "Remote ip Address",)
        cdp_detail.write("CDP_Nei_Info", "G", "1", "Platform",)
        cdp_detail.write("CDP_Nei_Info", "H", "1", "Native VLAN",)
        cdp_detail.filter_cols("CDP_Nei_Info", "A", "30")
        cdp_detail.filter_cols("CDP_Nei_Info", "B", "25")
        cdp_detail.filter_cols("CDP_Nei_Info", "C", "25")
        cdp_detail.filter_cols("CDP_Nei_Info", "D", "25")
        cdp_detail.filter_cols("CDP_Nei_Info", "E", "45")
        cdp_detail.filter_cols("CDP_Nei_Info", "F", "25")
        cdp_detail.filter_cols("CDP_Nei_Info", "G", "25")
        cdp_detail.filter_cols("CDP_Nei_Info", "H", "25")

        index = 2
        for entries in CDP_Info_List:
            cdp_detail.write("CDP_Nei_Info", "A", f"{index}", entries["Local Hostname"],)
            cdp_detail.write("CDP_Nei_Info", "B", f"{index}", entries["Local ip Address"],)
            cdp_detail.write("CDP_Nei_Info", "C", f"{index}", entries["Local Interface"],)
            cdp_detail.write("CDP_Nei_Info", "D", f"{index}", entries["Remote Interface"],)
            cdp_detail.write("CDP_Nei_Info", "E", f"{index}", entries["Remote Host"],)
            cdp_detail.write("CDP_Nei_Info", "F", f"{index}", entries["Remote ip Address"],)
            cdp_detail.write("CDP_Nei_Info", "G", f"{index}", entries["Platform"],)
            if "Native VLAN" in entries:
                cdp_detail.write("CDP_Nei_Info", "H", f"{index}", entries["Native VLAN"],)
            else:
                cdp_detail.write("CDP_Nei_Info", "H", f"{index}", "Not Found",)
            index += 1
    except Exception as err:
        error_log("Main Function Error: An unknown error occurred!")
        error_log(f"\t Error: {err}")
    finally:
        end = timer.time()
        elapsed = (end - start) / 60
        output_log(f"Total execution time: {elapsed:.3} minutes.",)
        output_log(f"Script Complete for site: {Sitecode}",)
        messagebox(f"Script Complete for site: {Sitecode}", "Script Complete")


if __name__ == "__main__":
    main()
