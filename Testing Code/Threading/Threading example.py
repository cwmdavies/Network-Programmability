import pandas as pd
from pandas import ExcelWriter
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading


# These capture errors relating to hitting ctrl+C (I forget the source)
#signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # IOError: Broken pipe
#signal.signal(signal.SIGINT, signal.SIG_DFL)  # KeyboardInterrupt: Ctrl-C

# Set the number of threads, I've found that 5 is the max
num_threads = 30

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()

# Define some variables to be used later in the script and ask user for some input
username = input('Enter Username: ')
#username = 'adm.cruickshank'
password = getpass.getpass('Enter Password: ')
# Get the name of site being audited
#site_name = input('Enter the name of the site or location being audited: ')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
#tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Last_Input_Check\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\CDP_Check\\'

# Define the current date and error log file info
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y_")  # Set timestamp to current system time
cdp_output = tftp_path_final + timestamp + '_CDP_Check.txt'
#spreadsheet_name = timestamp + 'Last_Input_Check_' + str(site_name).strip() + '.xlsx'

cdp_list = []

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)
#count_ip = df_read_ip.shape[0]  # Get the number of rows in column 1

# Create Spreadsheet with blank sheet
#writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name, engine='xlsxwriter')
#writer.save()
# Load the workbook
#workbook = load_workbook(tftp_path_final + spreadsheet_name)

# Start to iterate through the IP's in the file

def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()
        print("{}: Acquired IP: {}".format(i,ip_address))

        # Define a switch type
        switch = {
            "device_type": "cisco_ios",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
        }

        # Test the ssh connection and handle any errors and output to text file
        try:
            net_connect = ConnectHandler(**switch)
        except (AuthenticationException):
            with print_lock:
                print("\n{}: ERROR: Authenticaftion failed for {}. Stopping thread. \n".format(i,ip_address))
            q.task_done()
            continue 
        except (NetMikoTimeoutException):
            with print_lock:
                print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue
        except (SSHException):
            with print_lock:
                print("\n{}: SSH might not be enable on: {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue
        except (EOFError):
            with print_lock:
                print("\n{}: End of liner error attempting device: {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue

        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        net_connect.enable()
        switchname = net_connect.send_command ("sh ver | i uptime")  # Use to get the hostname
        switchname = switchname.split()[0]  # Get the first word which will be the switchname
        switchname = switchname.strip()  # Strip any trailing white space from the variable
        net_connect.send_config_set("ip tftp source-interface loopback 0")  # Set tftp source interface to loopback 0

        # Assign IOS commands to variables
        command_1 = "sh cdp nei"
        command_result = net_connect.send_command(command_1)
        cdp_list.append("\nCDP Neighbours for " + switchname + " :\n" + "\n" + command_result + "\n")

        #print(cdp_list)



        q.task_done()

def main():

    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    # Loop through the IP Address CSV and put the IP address into the queue
    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()

    
    #df1 = pd.DataFrame(data=[cdp_list],columns=["Result","col1","col2","col3"])
    
    #print(df1)
    #raise(SystemExit)


    cdp_write=open(cdp_output, "a")
    
    for val in cdp_list:
        cdp_write.write(val)
    
    cdp_write.close()
    print('*** CDP Check Completed ***')

if __name__ == "__main__":
    # Call the main function
    main()