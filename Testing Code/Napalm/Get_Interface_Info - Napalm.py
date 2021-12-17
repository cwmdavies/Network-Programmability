###############################################
#             Under Construction              #
#                Design Phase                 #
#                                             #
###############################################
#
#   A simple  napalm script that parses the output 
#   of the "show interface descriptions"
#   command and writes it in a neat format
#   to an excel spreadsheet.

import napalm
from openpyxl import load_workbook, Workbook
import datetime as time
from getpass import getpass
import ipaddress

debug = 1
username = input("Enter your Username: ")
password = getpass("Enter your Password: ")


def ip_check(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False


#######################################################################################################################
#          Logging Functions
#

def error_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{datetime} - {message}")
    error_file.write("\n")
    error_file.close()
    if debug == 1:
        print(message)


def output_log(message,):
    date_time_object = time.datetime.now()
    datetime = date_time_object.strftime("%d/%m/%Y %H:%M:%S")
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{datetime} - {message}")
    output_file.write("\n")
    output_file.close()
    if debug == 1:
        print(message)

#
#
#######################################################################################################################


def interfaces(ip):
    driver_ios = napalm.get_network_driver("ios")
    device = driver_ios(hostname=ip, username=username, password=password)
    device.open()
    device_interfaces = device.get_interfaces()
    device.close()
    return device_interfaces


def get_hostname(ip):
    driver_ios = napalm.get_network_driver("ios")
    device = driver_ios(hostname=ip, username=username, password=password)
    device.open()
    device_facts = device.get_facts()
    device.close()
    return device_facts["hostname"]


def main():
    ip_list = []
    with open("ips.txt", "r") as text:
        for IP_Address in text:
            ip_list.append(IP_Address.strip())

    for IP in ip_list:
        device_hostname = get_hostname(IP)
        device_interfaces = interfaces(IP)

        filename = f"Interfaces_{device_hostname}.xlsx"

        workbook = Workbook()
        workbook.save(filename=filename)
        workbook = load_workbook(filename=filename)
        workbook.create_sheet("Interface configuration")
        del workbook["Sheet"]
        ws = workbook["Interface configuration"]
        ws[f"A1"] = device_hostname
        ws[f"B1"] = IP
        ws[f"A3"] = "Interface"
        ws[f"B3"] = "is_enabled"
        ws[f"C3"] = "is_up"
        ws[f"D3"] = "Description"
        ws[f"E3"] = "MTU"
        ws[f"F3"] = "Speed"
        ws.column_dimensions['A'].width = "25"
        ws.column_dimensions['B'].width = "15"
        ws.column_dimensions['C'].width = "10"
        ws.column_dimensions['D'].width = "60"
        ws.column_dimensions['E'].width = "10"
        ws.column_dimensions['F'].width = "10"

        index = 4
        for interface in device_interfaces:
            ws[f"A{index}"] = interface
            ws[f"B{index}"] = device_interfaces[interface]["is_enabled"]
            ws[f"C{index}"] = device_interfaces[interface]["is_up"]
            ws[f"D{index}"] = device_interfaces[interface]["description"]
            ws[f"E{index}"] = device_interfaces[interface]["mtu"]
            ws[f"F{index}"] = device_interfaces[interface]["speed"]
            workbook.save(filename=filename)
            index += 1


if __name__ == "__main__":
    main()
