###############################################
#                                             #
#             Under Contruction               #
#                                             #
###############################################

import os
import re
import time
from multiprocessing.pool import ThreadPool
import paramiko
from datetime import datetime
from openpyxl import load_workbook, Workbook
from getpass import getpass

username = input("Type in your username: ")
password = getpass(prompt="Type in your password: ")
default_domain_name = "cns.muellergroup.com"
port = "22"

IP_list = []
hostname_list = []
fqdn_list = []
matched_list = []

class __excel:
    def __init__(self, name):
        self.i = 0
        self.name = name
        self.filename = self.name + ".xlsx"
        workbook = Workbook()
        workbook.save(filename=self.filename)
    def get_sheets(self):
        workbook = load_workbook(filename=self.filename)
        return workbook.sheetnames
    def add_sheets(self, *col_name):
        workbook = load_workbook(filename=self.filename)
        for value in col_name:
            if value not in workbook.sheetnames:
                col_name = workbook.create_sheet(value, self.i)
                self.i += 1
            else:
                output_log(f"{value} already exists in {self.name}. Ignoring column creation!")
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(filename=self.filename)
    def write(self, sheet, key, index, value):
        workbook = load_workbook(filename=self.filename)
        ws = workbook[f"{sheet}"]
        ws[f"{key}{index}"] = value
        workbook.save(filename=self.filename)

def error_log(message):
    dateTimeObj = datetime.now()
    print(f"{message}")
    error_file = open("Error Log.txt", "a")
    error_file.write(f"{dateTimeObj} - {message}")
    error_file.close()

def output_log(message):
    dateTimeObj = datetime.now()
    print(f"{message}")
    output_file = open("Output Log.txt", "a")
    output_file.write(f"{dateTimeObj} - {message}")
    output_file.close()

def open_session(IP):
    try:
        output_log(f"Connected to: {IP}")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=IP, port=port, username=username, password=password)
        return ssh, True
    except paramiko.ssh_exception.AuthenticationException:
        error_log(f"Authentication to IP: {IP} failed! Please check your IP, username and password.\n")
        return None, False
    except paramiko.ssh_exception.NoValidConnectionsError:
        error_log(f"Unable to connect to IP: {IP}!\n")
        return None, False
    except (ConnectionError, TimeoutError):
        error_log(f"Timeout error occured for IP: {IP}!\n")
        return None, False

def extract_cdp_neighbors(IP):
    interface_names = []
    command = "sh cdp neighbors | exclude (SEP|AIR)"
    regex = r"^.{17}(\b(Ten|Gig|Loo|Vla|F).{15})"
    ssh, connection = open_session(IP)
    if not connection:
        return None
    try:
        _, output, _ = ssh.exec_command(command)
        output = output.read()
        output = output.decode("utf-8")
        matches = re.finditer(regex, output, re.MULTILINE)
        for match in matches:
            temp_interface_name = match.group(1)
            temp_interface_name = temp_interface_name.strip()
            interface_names.append(temp_interface_name)
        return interface_names
    except paramiko.ssh_exception.SSHException:
        error_log(f"Extract CDP Neighbor Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
        return None
    finally:
        ssh.close()

def neighbor_detail(IP, commands):
    formatted_commands = []
    global IP_list
    regex = r"(?=[\n\r].*IP address:[\s]*([^\n\r]*))"
    ssh, connection = open_session(IP)
    if not connection:
        return None
    try:
        channel = ssh.invoke_shell()
        stdin = channel.makefile("wb")
        output = channel.makefile("rb")
        formatted_commands.append("'''")
        for c in commands:
            formatted_commands.append(c)
        formatted_commands.append("'''")
        formatted_commands = "\n".join(formatted_commands)
        stdin.write(str.encode(formatted_commands))
        output = output.read()
        output = output.decode("utf-8")
        stdin.close()
        matches = re.finditer(regex, output, re.MULTILINE)
        i = 1
        for match in matches:
            if match.group(i):
                found_IP = match.group(i)
                if found_IP not in IP_list:
                    IP_list.append(found_IP)
    except paramiko.ssh_exception.SSHException:
        error_log(f"Neighbor Detail Function Error: There is an error connecting or establishing SSH session to IP Address {IP}")
    finally:
        ssh.close()

def find_IPs(IP):
    commands = []
    interface_names = extract_cdp_neighbors(IP)
    if not interface_names:
        return -1
    for name in interface_names:
        commands.append(f"show cdp neighbors {name} detail | include IP")
    commands.append("exit")
    neighbor_detail(IP, commands)

def get_hostname_and_domain_name(IP):
    hostname = None
    domain_name = default_domain_name
    regex_hostname = r"^\bhostname[\s\r]+(.*)$"
    regex_domain_name = r"^IP[\s\r]domain-name[\s\r]+(.*)$"
    ssh, connection = open_session(IP)
    if not connection:
        return "-1", default_domain_name
    try:
        channel = ssh.invoke_shell()
        stdin = channel.makefile("wb")
        output = channel.makefile("rb")
        stdin.write(
            """
        show run | i hostname
        show run | i domain-name
        exit
        """
        )

        output = output.read()
        output = output.decode("utf-8").splitlines()
        output = "\n".join(output)
        hostname_matches = re.finditer(regex_hostname, output, re.MULTILINE)
        for h in hostname_matches:
            hostname = h.group(1)

        domain_name_matches = re.finditer(regex_domain_name, output, re.MULTILINE)
        for d in domain_name_matches:
            domain_name = d.group(1)

        stdin.close()
        return hostname, domain_name
    except paramiko.ssh_exception.SSHException:
        error_log(f"There is an error connecting or establishing SSH session to IP Address {IP}")
    finally:
        ssh.close()


def match_name_with_IP_address(IP, hostname, domain_name):
    temp_data = []
    command = "show IP interface brief | exclude unassigned"
    regex = r"(^[GTVLF].{22})+(.{16})"
    ssh, connection = open_session(IP)
    if not connection:
        return None
    try:
        _, stdout, _ = ssh.exec_command(command)
        stdout = stdout.read()
        stdout = stdout.decode("utf-8").splitlines()
        stdout = "\n".join(stdout)
        matches = re.finditer(regex, stdout, re.MULTILINE)
        for match in matches:
            temp_interface = match.group(1)
            temp_interface = temp_interface.strip()
            temp_IP = match.group(2)
            temp_IP = temp_IP.strip()
            shortened = temp_interface[0:4]
            temp_no = []
            for j in range(1, len(temp_interface)):
                if temp_interface[-j] == "/" or temp_interface[-j].isdigit():
                    temp_no.append(temp_interface[-j])
            temp_name = shortened + "".join(temp_no[::-1])
            name = [temp_name.upper(),temp_IP,hostname,domain_name]
            temp_data.append(name)
        return temp_data
    except paramiko.ssh_exception.SSHException:
        error_log(f"There is an error connecting or establishing SSH session to IP Address {IP}")
    finally:
        ssh.close()

def write_file(IP):
    global fqdn_list
    hostname, domain_name = get_hostname_and_domain_name(IP)
    if hostname == "-1":
        return -1
    elif not hostname:
        error_log(f"Hostname for IP Address: {IP} couldn't be found!")
        return -2
    elif hostname not in hostname_list:
        hostname_list.append(hostname)
        output_log(f"Hostname: {hostname}")
        fqdn = f"{hostname}.{domain_name}"
        fqdn_list.append(fqdn)
        lines_to_write = match_name_with_IP_address(IP, hostname, domain_name)
        for line in lines_to_write:
            matched_list.append(line)
    else:
        output_log(f"Hostname: {hostname} is in the list of hostnames")
        return -3

def main():
    global IP_list
    start = time.time()
    
    CDP_Recursion = __excel("CDP Recursion")
    CDP_Recursion.add_sheets("Found IPs","FQDN","DNS",)
    CDP_Recursion.write("DNS","A","1","Interface",)
    CDP_Recursion.write("DNS","B","1","IP Address",)
    CDP_Recursion.write("DNS","C","1","Hostname",)
    CDP_Recursion.write("DNS","D","1","Domain Name",)

    with open("IP.txt") as f:
        IP = f.readline()
    IP_list.append(IP)

    pool = ThreadPool(15)
    i = 0

    while i < len(IP_list) < 15:
        find_IPs(IP_list[i])
        i = i + 1

    while i < len(IP_list):
        limit = i + min(15, (len(IP_list) - i))
        hostnames = IP_list[i:limit]
        pool.map(find_IPs, hostnames)
        i = limit

    pool.map(write_file, IP_list)
    pool.close()
    pool.join()

    IP_cellnumber = 1
    for IP in IP_list:
        CDP_Recursion.write("Found IPs","A",f"{IP_cellnumber}",f"{IP}")
        IP_cellnumber += 1
    
    FQDN_cellnumber = 1
    for fqdn in fqdn_list:
        CDP_Recursion.write("FQDN","A",f"{FQDN_cellnumber}",f"{fqdn}")
        FQDN_cellnumber += 1

    DNS_cellnumber = 2
    for a,b,c,d in matched_list:
        CDP_Recursion.write("DNS","A",f"{DNS_cellnumber}",f"{a}",)
        CDP_Recursion.write("DNS","B",f"{DNS_cellnumber}",f"{b}",)
        CDP_Recursion.write("DNS","C",f"{DNS_cellnumber}",f"{c}",)
        CDP_Recursion.write("DNS","D",f"{DNS_cellnumber}",f"{d}",)
        DNS_cellnumber += 1
    
    end = time.time()
    elapsed = (end - start) / 60
    output_log(f"Total execution time: {elapsed:.7} minutes.")

if __name__ == "__main__":
    main()